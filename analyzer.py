import streamlit as st
import tempfile
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import numpy as np
import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
from sklearn.mixture import GaussianMixture
from sklearn.preprocessing import StandardScaler
from scipy.interpolate import make_interp_spline
from io import BytesIO
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objs as go
import plotly.express as px
from PIL import Image as PILImage
import os
import uuid

def select_filters(data, filters, num_filters):
    filters_dict = {}
    st.session_state.show_data_filter = {}
    if 'filter_tabs' not in st.session_state:
        st.session_state.filter_tabs = []
    st.session_state.filter_tabs = [f"Analysis Task {i+1}" for i in range(num_filters)]
    tabs = st.tabs(st.session_state.filter_tabs)
    for i in range(num_filters):
        with tabs[i]:
            st.subheader(f"Select Filters for Analysis Task {i + 1}")
            options_dict = {}
            st.session_state.show_data_filter[i] = {}
            for j in ["Target_Brand", "Competitor_Brand"]:
                st.session_state["data_copy_filter"]=data
                options_dict[j] = {}
                if j=='Target_Brand':
                    key_display = 'P&G'
                else:
                    key_display = 'Competitor'
                st.write(f"**{key_display}:**")
                st.session_state.show_data_filter[i][key_display]= data
                for k in filters:
                    data_copy_filter = st.session_state["data_copy_filter"]
                    option = st.selectbox(f"{key_display} - For the task analysis {i + 1}, select the filter for column {k}", data_copy_filter[k].unique())
                    data_copy_filter = data_copy_filter[data_copy_filter[k]==option]
                    st.session_state["data_copy_filter"] = data_copy_filter
                    st.session_state.show_data_filter[i][key_display] = data_copy_filter
                    try:
                        with st.expander(f"Task Analysis {i+1}: View Current {key_display} Avg Volume Share based on applied filters for {k}", expanded=False):
                            data_1 = st.session_state.show_data_filter[i][key_display].groupby(filters[filters.index(k)+1:filters.index(k)+2]+['Week End Date'])['Vol Share'].sum().reset_index()
                            data_2 = data_1.groupby(filters[filters.index(k)+1:filters.index(k)+2])['Vol Share'].mean().reset_index().sort_values(by='Vol Share', ascending=False)
                            data_2['Vol Share % within selection'] = (data_2['Vol Share']/sum(data_2['Vol Share']))*100
                            st.dataframe(data_2)
                    except:
                        pass
                    options_dict[j][k] = option
            filters_dict[i] = options_dict
    return filters_dict

@st.cache_data
def filter(data, filters_dict, num_filters, key_column):
    data_dict = {}
    meta_dict = {}
    for i in range(num_filters):
        options_dict = filters_dict[i]  # Access the correct options_dict level
        filter_data = {}
        for j in ["Target_Brand", "Competitor_Brand"]:
            query = " and ".join([f"`{k}` == '{options_dict[j][k]}'" for k in options_dict[j].keys()])
            data_subset = data.query(query)
            data_subset = data_subset.groupby("Week End Date").agg({"Vol Share": "sum", "Average Unit Price": "mean"}).reset_index()
            data_subset.columns = ["Week End Date", f"Vol_Share_{j}", f"Average_Unit_Price_{j}"]
            filter_data[j] = data_subset
        if filter_data["Target_Brand"].empty or filter_data["Competitor_Brand"].empty:
            continue
        data_merged = pd.merge(filter_data["Target_Brand"], filter_data["Competitor_Brand"], on="Week End Date", how="left")
        data_merged["Target_Brand Price ix"] = (data_merged["Average_Unit_Price_Target_Brand"] / data_merged["Average_Unit_Price_Competitor_Brand"]) * 100
        data_merged["Competitor_Brand Price ix"] = (data_merged["Average_Unit_Price_Competitor_Brand"] / data_merged["Average_Unit_Price_Target_Brand"]) * 100
        data_merged["Vol_Share_Total"] = data_merged["Vol_Share_Target_Brand"] + data_merged["Vol_Share_Competitor_Brand"]
        data_merged.fillna(0,inplace=True)
        key = f"{options_dict['Target_Brand']['Brand'][:3]}_{options_dict['Target_Brand'][key_column]}_{options_dict['Competitor_Brand']['Brand'][:3]}_{options_dict['Competitor_Brand'][key_column]}"
        data_dict[key] = data_merged
        meta_dict[key] = options_dict
    return data_dict, meta_dict

@st.cache_resource
def view_data(data_dict, meta_dict):
    for key in data_dict.keys():
        st.markdown(f"<h3>{key}</h3>", unsafe_allow_html=True)
        st.write("Target Brand Filters:")
        for k in meta_dict[key]["Target_Brand"].keys():
            st.write(f"{k} : {meta_dict[key]['Target_Brand'][k]}")
        st.write("Competitor Brand Filters:")
        for k in meta_dict[key]["Competitor_Brand"].keys():
            st.write(f"{k} : {meta_dict[key]['Competitor_Brand'][k]}")
        st.write("Data:")
        st.dataframe(data_dict[key])
        st.markdown("----")
    
@st.cache_resource        
def _add_data_worksheet(data_dict, meta_dict):
    wb = Workbook()
    for sheetname, df in data_dict.items():
        ws = wb.create_sheet(title=sheetname)
        ws.append(df.columns.tolist())
        for row in df.itertuples(index=False, name=None):
            ws.append(row)
        ws["N1"].value = "Metadata:"
        ws["N1"].font = openpyxl.styles.Font(bold=True)
        ws["N1"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws["N1"].fill = openpyxl.styles.PatternFill("solid", fgColor="00b0f0")
        ws["N3"].value = "Target Brand:"
        ws["N3"].font = openpyxl.styles.Font(bold=True)
        ws["N3"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws["P3"].value = "Competitor Brand:"
        ws["P3"].font = openpyxl.styles.Font(bold=True)
        ws["P3"].alignment = openpyxl.styles.Alignment(horizontal="center")
        p = 4
        for k in meta_dict[sheetname]["Target_Brand"].keys():
            ws[f"N{p}"].value = k
            ws[f"O{p}"].value = meta_dict[sheetname]["Target_Brand"][k]
            ws[f"P{p}"].value = k
            ws[f"Q{p}"].value = meta_dict[sheetname]["Competitor_Brand"][k]
            p+=1
            
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    return wb

def generate_charts(wb, x_col_index, y_col_index_1, y_col_index_2, config_dict, header, header_index, start_index, limit_dict):
    for name in wb.sheetnames:
        ws = wb[name]
        xvalues = Reference(ws, min_col=y_col_index_2, min_row=x_col_index, max_row=config_dict[name])
        yvalues = Reference(ws, min_col=y_col_index_1, min_row=x_col_index, max_row=config_dict[name])
        
        scatter_chart = ScatterChart()
        scatter_chart.title = "Price Index vs Share of Volume"
        scatter_chart.x_axis.title = 'Price Index'
        scatter_chart.y_axis.title = 'Share of Volume'
        scatter_chart.scatterStyle = None
        
        series = Series(yvalues, xvalues, title="Price Index vs Share of Volume")
        scatter_chart.series.append(series)
        scatter_chart.x_axis.major_gridlines = None
        scatter_chart.y_axis.major_gridlines = None
        scatter_chart.x_axis.scaling.min = limit_dict[name]['x'][0]
        scatter_chart.x_axis.scaling.max = limit_dict[name]['x'][1]
        scatter_chart.y_axis.scaling.min = limit_dict[name]['y'][0]
        scatter_chart.y_axis.scaling.max = limit_dict[name]['y'][1]
        scatter_chart.x_axis.number_format = '0'
        scatter_chart.legend = None

        trendline = Trendline()
        trendline.type = "linear" 
        trendline.spPr = GraphicalProperties(solidFill=ColorChoice(prstClr="orange"))
        series.trendline = trendline
        
        ws[header_index].value = header
        ws[header_index].font = openpyxl.styles.Font(bold=True)
        ws.add_chart(scatter_chart, start_index)
    return wb
        
def create_scatter_plot(df, x_column, y_column, x_lable, y_label, title):
    # Scatter plot using Plotly Express
    fig = px.scatter(df, x=x_column, y=y_column, 
                     labels={x_column: x_lable, y_column: y_label},
                     title=title)
    
    x_numeric = df[x_column].values
    z = np.polyfit(x_numeric, df[y_column], 1)
    p = np.poly1d(z)

    fig.add_scatter(x=df[x_column], y=p(x_numeric), mode='lines', name='Trendline', line=dict(color='red'))

    return fig

def display_chart(data):
    for key in data.keys():
        img = create_scatter_plot(data[key], "Target_Brand Price ix", "Vol_Share_Target_Brand", "Price Index", "Target Brand Vol Share", "Target Brand Vol Share vs Price Index")
        st.markdown(f"<h3>{key}</h3>", unsafe_allow_html=True)
        st.plotly_chart(img, caption='Price Index vs Target Brand Volume Share', use_column_width=True)
        img1 = create_scatter_plot(data[key], "Target_Brand Price ix", "Vol_Share_Total", "Price Index", "Total Vol Share", "Total Vol Share vs Price Index")
        st.plotly_chart(img1, caption='Price Index vs Total Volume Share', use_column_width=True)
        st.markdown("----")
        
@st.cache_resource
def generate_stats(data_dict, x_column, y_column):
    stats_dict = {}
    for key in data_dict.keys():
        stats_dict[key] = {}
        data_filtered = data_dict[key]
        x_values = data_filtered[x_column].values
        y_values = data_filtered[y_column].values
        model = LinearRegression()
        model.fit(x_values.reshape(-1, 1), y_values)
        y_pred = model.predict(x_values.reshape(-1, 1))
        r2 = r2_score(y_values, y_pred)
        correl = data_filtered[x_column].corr(data_filtered[y_column])
        stats_dict[key]["Coefficient"] = model.coef_[0]
        stats_dict[key]["Intercept"] = model.intercept_
        stats_dict[key]["r2"] = r2
        stats_dict[key]["correl"] = correl
    return stats_dict


def write_stats(wb, stats_dict, header, header_start, start_column_label, start_column_value, start_index):
    for name in wb.sheetnames:
        ws = wb[name]
        ws[header_start].value = header
        ws[header_start].font = openpyxl.styles.Font(bold=True)
        ws[f"{start_column_label}{start_index}"].value = "Coefficient"
        ws[f"{start_column_value}{start_index}"].value = round(stats_dict[name]["Coefficient"],3)
        ws[f"{start_column_label}{start_index+1}"].value = "Intercept"
        ws[f"{start_column_value}{start_index+1}"].value = round(stats_dict[name]["Intercept"],2)
        ws[f"{start_column_label}{start_index+2}"].value = "r2"
        ws[f"{start_column_value}{start_index+2}"].value = round(stats_dict[name]["r2"],2)
        ws[f"{start_column_label}{start_index+3}"].value = "correl"
        ws[f"{start_column_value}{start_index+3}"].value = round(stats_dict[name]["correl"],2)
        ws[f"{start_column_label}{start_index+4}"].value = "Equation"
        ws[f"{start_column_value}{start_index+4}"].value = f"y = {stats_dict[name]['Coefficient']}x + {stats_dict[name]['Intercept']}"
    return wb

def display_stats(stats, stats_1):
    for key in stats.keys():
        st.markdown(f"<h3>{key}</h3>", unsafe_allow_html=True)
        st.write(f"**Target Volume Share vs Price Index for {key}:**")
        st.markdown(f"Coefficient: {round(stats[key]['Coefficient'],3)}")
        st.markdown(f"Intercept: {round(stats[key]['Intercept'],2)}")
        st.markdown(f"r2: {round(stats[key]['r2'],2)}")
        st.markdown(f"correl: {round(stats[key]['correl'],2)}")
        st.markdown(f"Equation: y = {stats[key]['Coefficient']}x + {stats[key]['Intercept']}")
        st.write(f"**Total Volume Share vs Price Index for {key}:**")
        st.markdown(f"Coefficient: {round(stats_1[key]['Coefficient'],3)}")
        st.markdown(f"Intercept: {round(stats_1[key]['Intercept'],2)}")
        st.markdown(f"r2: {round(stats_1[key]['r2'],2)}")
        st.markdown(f"correl: {round(stats_1[key]['correl'],2)}")
        st.markdown(f"Equation: y = {round(stats_1[key]['Coefficient'],3)}x + {round(stats_1[key]['Intercept'],3)}")
        st.markdown("----")

# @st.cache_data
def gmm_clustering_optimal(data_dict, cols):
    data_gmm = {}
    summary_gmm = {}
    optimal_clusters_gmm = {}
    for key in data_dict.keys():
        data_select = data_dict[key]
        scaler = StandardScaler()
        df_scaled = scaler.fit_transform(data_select[cols])
        
        n_components_range = range(1, 10)
        bic_scores = []
        aic_scores = []
        
        for n in n_components_range:
            gmm = GaussianMixture(n_components=n, random_state=42)
            gmm.fit(df_scaled)
            
            bic_scores.append(gmm.bic(df_scaled))  # BIC score
            aic_scores.append(gmm.aic(df_scaled))  # AIC score
        optimal_n_components = n_components_range[np.argmin(bic_scores)]  # Choose based on BIC (or AIC)
        gmm_optimal = GaussianMixture(n_components=optimal_n_components, random_state=42)
        data_select['price_corridor'] = gmm_optimal.fit_predict(df_scaled)
        corridor_summary = data_select.groupby('price_corridor').agg(
            count_points =('price_corridor', 'count'),
            avg_price=('Target_Brand Price ix', 'mean'),
            avg_volume=('Vol_Share_Target_Brand', 'mean'),
            min_price=('Target_Brand Price ix', 'min'),
            max_price=('Target_Brand Price ix', 'max'),
            min_volume=('Vol_Share_Target_Brand', 'min'),
            max_volume=('Vol_Share_Target_Brand', 'max')
        ).reset_index()
        
        correlations = []
        slopes=[]
        for corridor in corridor_summary['price_corridor']:
            corridor_data = data_select[data_select['price_corridor'] == corridor]
            correlation = corridor_data['Target_Brand Price ix'].corr(corridor_data['Vol_Share_Target_Brand'])
            correlations.append(correlation)
            X = corridor_data[['Target_Brand Price ix']]
            y = corridor_data['Vol_Share_Target_Brand']
            reg = LinearRegression().fit(X, y)
            slope = reg.coef_[0]
            slopes.append(slope)
        corridor_summary['correlation'] = correlations
        corridor_summary['slope'] = slopes
        
        data_gmm[key] = data_select
        summary_gmm[key] = corridor_summary
        optimal_clusters_gmm[key] = optimal_n_components
        
    return data_gmm, summary_gmm, optimal_clusters_gmm

def gmm_clustering_custom(data_dict, cols, num_clusters):
    data_gmm = {}
    summary_gmm = {}
    for key in data_dict.keys():
        data_select = data_dict[key]
        scaler = StandardScaler()
        df_scaled = scaler.fit_transform(data_select[cols])
        
        gmm_optimal = GaussianMixture(n_components=int(num_clusters), random_state=42)
        data_select['price_corridor'] = gmm_optimal.fit_predict(df_scaled)
        corridor_summary = data_select.groupby('price_corridor').agg(
            count_points =('price_corridor', 'count'),
            avg_price=('Target_Brand Price ix', 'mean'),
            avg_volume=('Vol_Share_Target_Brand', 'mean'),
            min_price=('Target_Brand Price ix', 'min'),
            max_price=('Target_Brand Price ix', 'max'),
            min_volume=('Vol_Share_Target_Brand', 'min'),
            max_volume=('Vol_Share_Target_Brand', 'max')
        ).reset_index()
        
        correlations = []
        slopes=[]
        for corridor in corridor_summary['price_corridor']:
            corridor_data = data_select[data_select['price_corridor'] == corridor]
            correlation = corridor_data['Target_Brand Price ix'].corr(corridor_data['Vol_Share_Target_Brand'])
            correlations.append(correlation)
            X = corridor_data[['Target_Brand Price ix']]
            y = corridor_data['Vol_Share_Target_Brand']
            reg = LinearRegression().fit(X, y)
            slope = reg.coef_[0]
            slopes.append(slope)
        corridor_summary['correlation'] = correlations
        corridor_summary['slope'] = slopes
        
        data_gmm[key] = data_select
        summary_gmm[key] = corridor_summary
    return data_gmm, summary_gmm

def plot_corridor(data_dict):
    plot_dict = {}
    colors = ['red', 'green', 'blue', 'purple', 'orange', 'yellow', 'cyan', 'magenta', 'brown', 
              'pink', 'lime', 'darkblue', 'lightgray', 'darkgreen']

    for key in data_dict.keys():
        data_plot = data_dict[key]
        fig = go.Figure()
        unique_corridors = data_plot['price_corridor'].nunique()
        
        for corridor in range(unique_corridors):
            subset = data_plot[data_plot['price_corridor'] == corridor]
            fig.add_trace(go.Scatter(
                x=subset['Target_Brand Price ix'],
                y=subset['Vol_Share_Target_Brand'],
                mode='markers',
                marker=dict(color=colors[corridor]),
                name=f'Corridor {corridor}'
            ))
        
        fig.update_layout(
            title='Price Corridors',
            xaxis_title='Price Index',
            yaxis_title='Target Volume Share',
            showlegend=True
        )
        plot_dict[key] = fig
    return plot_dict

def write_corridor_summary_streamlit(corridor_summary, plots_dict, optimal_clusters={}):
    for key in corridor_summary.keys():
        st.markdown(f"<h3>{key} :</h3>", unsafe_allow_html=True)
        if optimal_clusters!={}:
            st.markdown(f"Optimal number of clusters for {key} is {optimal_clusters[key]}")
        st.write(f"Price Corridor Summary for {key}")
        st.dataframe(corridor_summary[key])
        st.write(f"Price Corridor Plot for {key}")
        st.plotly_chart(plots_dict[key], use_container_width=True)
        corridor_summary_analyze = corridor_summary[key].copy()
        corridor_summary_analyze.sort_values(by='correlation',inplace=True)
        idx = corridor_summary_analyze.loc[(corridor_summary_analyze["correlation"]<=-0.7),"price_corridor"].unique()
        idx1 = corridor_summary_analyze.loc[(corridor_summary_analyze["correlation"]<=-0.4)&(corridor_summary_analyze["correlation"]>-0.7),"price_corridor"].unique()
        if idx.any() or idx1.any():
            st.markdown("**Observations :**")
            for id in idx:
                st.write(f"The corridor no {id} has a high impact on the volume share")
            for id in idx1:
                st.write(f"The corridor no {id} has a decent impact on the volume share")
        st.markdown("----")

def append_data_in_rows_openpyxl(wb, sheet_name, data_work, start_row, start_col):
    ws = wb[sheet_name]
    for col_idx, column_name in enumerate(data_work.columns, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=column_name)

    for row_idx, row in enumerate(data_work.itertuples(index=False, name=None), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=value)
            
    return wb
            
def save_plot_as_image(fig):
    img_bytes = fig.to_image(format="png")
    img_stream = BytesIO(img_bytes)
    img_stream.seek(0)
    return img_stream

def insert_image_into_excel(img_stream, wb, name, position):
    ws = wb[name]
    unique_filename = f"{uuid.uuid4()}.png"
    with tempfile.NamedTemporaryFile(suffix=".png", name=unique_filename, delete=False) as tmp_file:
        tmp_file.write(img_stream.read()) 
        tmp_file.flush() 
        img_for_excel = Image(tmp_file.name)
    ws.add_image(img_for_excel, position)
    
    return wb
def add_img_bytes_to_sheet(img_stream, wb, name, cell) -> None:
    ws = wb[name]
    img = PILImage.open(img_stream)
    image = openpyxl.drawing.image.Image(img)
    def data_fillin(imgb):
        bobj = img_stream.getbuffer().tobytes()
        def wrap():
            
            return bobj
        return wrap
    image._data = data_fillin(img_stream)
    image.anchor = cell
    ws.add_image(image, cell)
    
    return wb
            
def write_corridor_summary_excel(wb, corridor_summary_optimal, corridor_summary_custom, plots_dict_optimal, plots_dict_custom):
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=35, max_col=50):
            for cell in row:
                cell.value = None
        ws["AI1"].value = "Price Corridor Summary"
        ws["AI1"].font = openpyxl.styles.Font(bold=True)
        ws["AI1"].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws["AI1"].fill = openpyxl.styles.PatternFill("solid", fgColor="00b0f0")
        ws["AI3"].value = "Price Corridor Summary"
        ws["AI3"].font = openpyxl.styles.Font(bold=True)
        ws["AI4"].value = "Optimal" 
        ws["AI4"].font = openpyxl.styles.Font(bold=True)
        wb = append_data_in_rows_openpyxl(wb, name, corridor_summary_optimal[name], 5, 35)
        ws = wb[name]
        ws["AU4"].value = "Custom" 
        ws["AU4"].font = openpyxl.styles.Font(bold=True)
        wb = append_data_in_rows_openpyxl(wb, name, corridor_summary_custom[name], 5, 47)
        # img_stream = save_plot_as_image(plots_dict_optimal[name])
        # wb = insert_image_into_excel(img_stream, wb, name,"AI25")
        # img_stream = save_plot_as_image(plots_dict_custom[name])
        # wb = insert_image_into_excel(img_stream, wb, name, "AS25")
    return wb

def main():
    st.set_page_config(page_title="Price Analytics",
                       page_icon=":books:")
    st.header("Price Analyzer for Home Care")
    uploaded_data = st.sidebar.file_uploader("Upload the price data", accept_multiple_files=False, type=["xlsx", "csv"])
    if 'data' not in st.session_state or st.session_state['uploaded_file'] != uploaded_data:
        if uploaded_data is not None:
            data = pd.read_excel(uploaded_data) if uploaded_data.name.endswith(".xlsx") else pd.read_csv(uploaded_data)
            data['Week End Date'] = pd.to_datetime(data['Week End Date'])
            st.session_state['data'] = data
            st.session_state['uploaded_file'] = uploaded_data
        else:
            st.warning("Please upload the data to proceed.")
            return

    data = st.session_state['data']

    filters = st.sidebar.multiselect("Select the colums for filters"
                                     "The columns should be selected in hierarchical order", data.columns)
    if not filters:
        st.warning("Please select the columns for filtering")
        return
    num_filters = st.sidebar.number_input("Number of Analysis Tasks", min_value=1, max_value=10)
    key_column = st.sidebar.selectbox("Select the column to use as key", data.columns)
    choose_option = st.sidebar.radio("Select the column combination for clustering", ["Price Index and Volume Share", "Price Index Only"], key="auto")
    num_clusters = st.sidebar.number_input("Enter the number of clusters for manual clustering", min_value=1, max_value=15, value=1, step=1)
    if 'data_dict' not in st.session_state:
        st.session_state['data_dict'] = None
    if 'meta_dict' not in st.session_state:
        st.session_state['meta_dict'] = None
    if 'analyzed' not in st.session_state:
        st.session_state['analyzed'] = False
    if 'wb' not in st.session_state:
        st.session_state['wb'] = None
    if 'data_summary_optimal' not in st.session_state:
        st.session_state['data_summary_optimal'] = None
    if 'corridor_summary_optimal' not in st.session_state:
        st.session_state['corridor_summary_optimal'] = None
    if 'data_summary_custom' not in st.session_state:
        st.session_state['data_summary_custom'] = None
    if 'corridor_summary_custom' not in st.session_state:
        st.session_state['corridor_summary_custom'] = None
    if 'optimal_clusters' not in st.session_state:
        st.session_state['optimal_clusters'] = None
    if 'plots_dict_optimal' not in st.session_state:
        st.session_state['plots_dict_optimal'] = None
    if 'plots_dict_custom' not in st.session_state:
        st.session_state['plots_dict_custom'] = None
    # data_dict, meta_dict = filter(data, filters, num_filters, key_column)
    filters_dict = select_filters(data, filters, num_filters)
    data_dict, meta_dict = filter(data, filters_dict, num_filters, key_column)
    config_dict = {}
    if st.button("Analyze"):
        with st.spinner("Analyzing..."):
            st.session_state['data_dict'] = data_dict
            st.session_state['meta_dict'] = meta_dict
            st.session_state['analyzed'] = True
            for key in st.session_state["data_dict"].keys():
                config_dict[key] = len(st.session_state["data_dict"][key]) + 1
            if st.session_state['analyzed']:
                tab1, tab2, tab3, tab4 = st.tabs(["View Data", "View Scatter Plots", "View Statistics", "Price Corridors"])
                with tab1:
                    view_data(st.session_state.data_dict, meta_dict) 
                    wb1 = _add_data_worksheet(st.session_state.data_dict, meta_dict)
                    st.session_state.wb = wb1
                with tab2:
                    limit_dict = {}
                    for key in st.session_state.data_dict.keys():
                        limit_dict[key] = {}
                        limit_dict[key]['x'] = [(st.session_state.data_dict[key]['Target_Brand Price ix'].min())*0.9, (st.session_state.data_dict[key]['Target_Brand Price ix'].max())*1.1]
                        limit_dict[key]['y'] = [(st.session_state.data_dict[key]['Vol_Share_Target_Brand'].min())*0.9, (st.session_state.data_dict[key]['Vol_Share_Target_Brand'].max())*1.1]
                    wb2 = generate_charts(st.session_state.wb, 2, 2, 6, config_dict, "Target Brand Vol Share vs Price Index:", "N9", "N10", limit_dict)
                    st.session_state.wb = wb2
                    
                    limit_dict = {}
                    for key in st.session_state.data_dict.keys():
                        limit_dict[key] = {}
                        limit_dict[key]['x'] = [(st.session_state.data_dict[key]['Target_Brand Price ix'].min())*0.9, (st.session_state.data_dict[key]['Target_Brand Price ix'].max())*1.1]
                        limit_dict[key]['y'] = [(st.session_state.data_dict[key]['Vol_Share_Total'].min())*0.9, (st.session_state.data_dict[key]['Vol_Share_Total'].max())*1.1]
                    wb3 = generate_charts(st.session_state.wb, 2, 8, 6, config_dict, "Total Vol Share vs Price Index:", "X9", "X10", limit_dict)
                    st.session_state.wb = wb3
                    display_chart(st.session_state.data_dict)
                with tab3:
                    stats = generate_stats(st.session_state.data_dict, "Target_Brand Price ix", "Vol_Share_Target_Brand")
                    wb4 = write_stats(st.session_state.wb, stats, "Target Brand Stats", "N27", "N", "O", 28)
                    st.session_state.wb = wb4
                    stats_1 = generate_stats(st.session_state.data_dict, "Target_Brand Price ix", "Vol_Share_Total")
                    wb5 = write_stats(st.session_state.wb, stats_1, "Target Brand Stats", "X27", "X", "Y", 28)
                    st.session_state.wb = wb5
                    display_stats(stats, stats_1)
                with tab4:
                    tab41, tab42 = st.tabs(["Optimal Number of Clusters", "Custom No of Clusters"])
                    with tab41:
                        # choose_option = st.radio("Select the column combination for clustering", ["Price Index and Volume Share", "Price Index Only"], key="auto")
                        if choose_option == "Price Index and Volume Share":
                            data_summary_optimal, corridor_summary_optimal, optimal_clusters = gmm_clustering_optimal(st.session_state.data_dict, ["Target_Brand Price ix", "Vol_Share_Target_Brand"])
                            
                        else:
                            data_summary_optimal, corridor_summary_optimal, optimal_clusters = gmm_clustering_optimal(st.session_state.data_dict, ["Target_Brand Price ix"])
                        st.session_state['data_summary_optimal'] = data_summary_optimal
                        st.session_state['corridor_summary_optimal'] = corridor_summary_optimal
                        st.session_state['optimal_clusters'] = optimal_clusters
                        plots_dict = plot_corridor(st.session_state["data_summary_optimal"])
                        st.session_state['plots_dict_optimal'] = plots_dict
                        write_corridor_summary_streamlit(st.session_state['corridor_summary_optimal'], st.session_state['plots_dict_optimal'], st.session_state['optimal_clusters'])
                    
                    with tab42:
                        # choose_option = st.radio("Select the column combination for clustering", ["Price Index and Volume Share", "Price Index Only"], key="manual")
                        # num_clusters = st.number_input("Enter the number of clusters", min_value=1, max_value=15, value=1, step=1)
                        if choose_option == "Price Index and Volume Share":
                            data_summary_custom, corridor_summary_custom= gmm_clustering_custom(st.session_state.data_dict, ["Target_Brand Price ix", "Vol_Share_Target_Brand"], num_clusters)
                            
                        else:
                            data_summary_custom, corridor_summary_custom= gmm_clustering_custom(st.session_state.data_dict, ["Target_Brand Price ix"], num_clusters)
                        st.session_state['data_summary_custom'] = data_summary_custom
                        st.session_state['corridor_summary_custom'] = corridor_summary_custom
                        plots_dict = plot_corridor(st.session_state["data_summary_custom"])
                        st.session_state['plots_dict_custom'] = plots_dict
                        write_corridor_summary_streamlit(st.session_state['corridor_summary_custom'], st.session_state['plots_dict_custom'])
                    wb6 = write_corridor_summary_excel(st.session_state['wb'], st.session_state['corridor_summary_optimal'], st.session_state['corridor_summary_custom'], st.session_state['plots_dict_optimal'], st.session_state['plots_dict_custom'])
                    st.session_state.wb = wb6
                output = BytesIO()
                wb_final = st.session_state.wb
                wb_final.save(output)
                output.seek(0)
                st.download_button("Download Analysis", output, "Price_Analytics.xlsx", "xlsx")
                
if __name__ == "__main__":
    main()
